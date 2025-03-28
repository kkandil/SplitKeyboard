from openpyxl import load_workbook

def generate_header_file(excel_file, output_file):
    # Open the Excel workbook
    wb = load_workbook(excel_file, data_only=True)

    # Layers (Layer1, Layer2, Layer3)
    layers = ['Layer1', 'Layer2', 'Layer3']

    #Tables
    ArrayNames = ["keyMapingLeftPress", "keyMapingLeftHold", "keyMapingRightPress", "keyMapingRightHold"]
    Tableindices = [20, 32, 44, 56]
    ThumbArrayNames = [  "thumbMapingLeftPress","thumbMapingLeftHold","thumbMapingRightPress","thumbMapingRightHold"]
    # Output header file
    with open(output_file, 'w') as f:
        # Write initial part of the header file
        f.write("#ifndef KEYMAP_H\n")
        f.write("#define KEYMAP_H\n\n")
         
        f.write("#include \"HID-Project.h\"\n\n")
        f.write("#include \"SplitKeyboard_Cfg.h\"\n\n")
        f.write("typedef struct {\n")
        f.write("    KeyboardKeycode Keycode;\n")
        f.write("    bool isShifted;\n")
        f.write("    bool isAMacro;\n")
        #f.write("    bool isText;\n")
        #f.write("    String Text;\n")
        f.write("    KeyboardKeycode MacroKey1;\n")
        f.write("    KeyboardKeycode MacroKey2;\n")
        f.write("    KeyboardKeycode MacroKey3;\n")
        f.write("} strKeyCode;\n\n")

        f.write("typedef struct {\n")
        f.write("    strKeyCode LayerMap[3];\n")
        f.write("} strKeyMapping;\n\n") 
        
        for table_start_col, ArrayNames_x in zip(Tableindices, ArrayNames):
            f.write(f"strKeyMapping {ArrayNames_x}[KEYBORD_ROWS_COUNT][KEYBORD_COLS_COUNT] = {{\n")
            
            for row_start in range(3, 19, 6):  # We process each block of 6 rows
                f.write("    {\n")  
                for i in range(row_start, row_start + 6):  # Each set of 6 rows
                    f.write("       {{")

                    # Process each layer (Layer1, Layer2, Layer3)
                    for layer_name in layers:
                        sheet = wb[layer_name]

                        # Read table columns: Key, isText, Text, isMacro, Key1, Key2, Key3
                        key = sheet.cell(i,table_start_col).value
                        
                        # Now, let's handle multi-character columns correctly
                       # is_text = sheet.cell(i,table_start_col+1).value
                        #text    = sheet.cell(i,table_start_col+2).value
                        is_macro= sheet.cell(i,table_start_col+3).value
                        key1    = sheet.cell(i,table_start_col+4).value
                        key2    = sheet.cell(i,table_start_col+5).value
                        key3    = sheet.cell(i,table_start_col+6).value
                        is_shifted = sheet.cell(i,table_start_col+7).value

                        # Format and write each row in the struct format
                        #f.write(f"{{ {key}, {is_shifted}, {is_text}, {is_macro}, \"{text}\", {key1}, {key2}, {key3} }}")
                        f.write(f"{{ {key}, {is_shifted}, {is_macro}, {key1}, {key2}, {key3} }}")
                        if layer_name!='Layer3':
                            f.write(",")
                    f.write("}}")
                    if i!=row_start + 5:
                        f.write(",\n")
                    else:
                        f.write("\n")
                if(row_start==15):
                    f.write("    }\n};\n")
                else:
                    f.write("    },\n")

        # Process tables for each layer
        for table_start_col, ArrayNames_x in zip(Tableindices, ThumbArrayNames):
            f.write(f"strKeyMapping {ArrayNames_x}[KEYBORD_ROWS_COUNT] = {{\n")
            row_start = 21
            for i in range(row_start, row_start + 3):  # set of 3 rows
                f.write("       {{")

                # Process each layer (Layer1, Layer2, Layer3)
                for layer_name in layers:
                    sheet = wb[layer_name]

                    # Read table columns: Key, isText, Text, isMacro, Key1, Key2, Key3
                    key = sheet.cell(i,table_start_col).value
                    
                    # Now, let's handle multi-character columns correctly
                    #is_text = sheet.cell(i,table_start_col+1).value
                    #text    = sheet.cell(i,table_start_col+2).value
                    is_macro= sheet.cell(i,table_start_col+3).value
                    key1    = sheet.cell(i,table_start_col+4).value
                    key2    = sheet.cell(i,table_start_col+5).value
                    key3    = sheet.cell(i,table_start_col+6).value
                    is_shifted = sheet.cell(i,table_start_col+7).value

                    # Format and write each row in the struct format
                    #f.write(f"{{ {key}, {is_shifted}, {is_text}, {is_macro}, \"{text}\", {key1}, {key2}, {key3} }}")
                    f.write(f"{{ {key}, {is_shifted}, {is_macro}, {key1}, {key2}, {key3} }}")
                    if layer_name!='Layer3':
                        f.write(",")
                f.write("}}")
                if i!=row_start + 2:
                    f.write(",\n")
                else:
                    f.write("\n")
            f.write("    };\n")

    
        # Close the file
        f.write("#endif // KEYMAP_H\n")

if __name__ == "__main__":
    excel_file = "KeyMapingGenerator.xlsx"  # Change to your file name
    output_file = "..\\Firmware\\SplitKeyboard_Left\\KeyMapping.h"  # The name of the output header file
    generate_header_file(excel_file, output_file)
    print("Header file generated successfully.")
